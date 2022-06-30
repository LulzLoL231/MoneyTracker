# -*- coding: utf-8 -*-
#
#  MoneyTracker: cmds - Orders
#  Created by LulzLoL231 at 28/06/22
#
import json
import logging
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from vkbottle.bot import Message
from vkbottle.tools import DocMessagesUploader
from vkbottle import BaseStateGroup, BotBlueprint
from vkbottle.dispatch.rules.base import FuncRule, StateRule

from config import cfg
from database.types import Order
from keyboards import Keyboards as keys
from database.main import Database, DB_LOCK


log = logging.getLogger('MoneyTracker')
bp = BotBlueprint('orders')


class AddOrder(BaseStateGroup):
    NAME = 'name'
    PRICE = 'price'
    AGENT = 'agent'
    VERIFY = 'verify'


class OrderInfo(BaseStateGroup):
    UID = 'uid'


async def make_export_file(orders: list[Order]) -> BytesIO:
    '''Makes XLSX file and return it.

    Args:
        orders (list[Order]): Array of Orders.

    Returns:
        BytesIO: XLSX file.
    '''
    period = (
        orders[0].start_date,
        orders[::-1][0].start_date
    )
    wb = Workbook()
    ws = wb.active
    ws.title = f'с {str(period[0])} по {str(period[1])}'
    ws['A1'] = '№ Заказа'
    ws['A1'].font = Font(bold=True)
    ws['B1'] = 'Цель'
    ws['B1'].font = Font(bold=True)
    ws['C1'] = 'Цена'
    ws['C1'].font = Font(bold=True)
    ws['D1'] = 'Агент'
    ws['D1'].font = Font(bold=True)
    ws['E1'] = 'Дата начала'
    ws['E1'].font = Font(bold=True)
    ws['F1'] = 'Дата оплаты'
    ws['F1'].font = Font(bold=True)
    rid = 2
    for order in orders:
        ws[f'A{rid}'] = order.uid
        ws[f'B{rid}'] = order.name
        ws[f'C{rid}'] = order.price
        ws[f'D{rid}'] = order.agent.name
        ws[f'E{rid}'] = order.start_date
        if order.end_date:
            ws[f'F{rid}'] = order.end_date
        else:
            ws[f'F{rid}'] = 'Не оплачен'
        rid += 1
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def is_order_cmd(payload: str) -> bool:
    if payload:
        return json.loads(payload).get('command', '').startswith('order#')
    return False


def is_del_order_cmd(payload: str) -> bool:
    if payload:
        return json.loads(payload).get('command', '').startswith('del_order#')
    return False


def is_end_order_cmd(payload: str) -> bool:
    if payload:
        return json.loads(payload).get('command', '').startswith('end_order#')
    return False


@bp.on.private_message(FuncRule(lambda m: m.text.lower() in ['/orders', 'заказы']))
@bp.on.private_message(payload={'command': 'orders'})
async def orders(msg: Message):
    user = await msg.get_user()
    log.info(f'Called by {user.first_name} {user.last_name} ({user.id})')
    async with DB_LOCK:
        orders = await Database.get_orders()
    inprog_orders = list(filter(
        (lambda o: not o.end_date),
        orders
    ))
    if not inprog_orders:
        cnt = 'Все заказы выполнены. Добавьте новый!'
    elif not orders:
        cnt = 'Заказы отсутствуют. Добавьте новый!'
    else:
        cnt = 'Текущие заказы:\n\n'
        full_price = 0
        for ord in inprog_orders:
            cnt += f'{ord.get_short_str()}\n'
            full_price += ord.price
        cnt += f'\nИтого: {full_price} руб.'
    await msg.answer(cnt, keyboard=keys.orders(inprog_orders, bool(orders)))


@bp.on.private_message(payload={'command': 'add_order'})
async def start_add_order(msg: Message):
    user = await msg.get_user()
    log.info(f'Called by {user.first_name} {user.last_name} ({user.id})')
    await bp.state_dispenser.set(msg.peer_id, AddOrder.NAME)
    await msg.answer(
        'Добавление заказа\n\nВведите цель заказа.',
        keyboard=keys.back('orders')
    )


@bp.on.private_message(StateRule(AddOrder.NAME))
async def add_order_name(msg: Message):
    user = await msg.get_user()
    log.info(f'Called by {user.first_name} {user.last_name} ({user.id})')
    if msg.text.lower() == 'назад':
        await bp.state_dispenser.delete(msg.peer_id)
        log.info(f'User #{msg.peer_id} canceled adding order.')
        await orders(msg)
        return
    await bp.state_dispenser.set(msg.peer_id, AddOrder.PRICE, name=msg.text)
    await msg.answer(
        'Введите цену заказа.',
        keyboard=keys.back('orders')
    )


@bp.on.private_message(StateRule(AddOrder.PRICE))
async def add_order_price(msg: Message):
    user = await msg.get_user()
    log.info(f'Called by {user.first_name} {user.last_name} ({user.id})')
    if msg.text.lower() == 'назад':
        await bp.state_dispenser.delete(msg.peer_id)
        log.info(f'User #{msg.peer_id} canceled adding order.')
        await orders(msg)
        return
    elif not msg.text.isdigit():
        await msg.answer(
            'ОШИБКА: В качестве цены должно выступать цифровое значение!\nПовторите ввод.',
            keyboard=keys.back('orders')
        )
        return
    async with DB_LOCK:
        agents = await Database.get_agents()
    if agents:
        await bp.state_dispenser.set(
            msg.peer_id, AddOrder.AGENT, price=int(msg.text),
            name=msg.state_peer.payload.get('name')
        )
        cnt = 'Выберите агента'
        key = keys.add_order_agents(agents)
    else:
        await bp.state_dispenser.delete(msg.peer_id)
        cnt = 'Нету доступных агентов. Сначала добавьте хотя-бы одного.'
        key = keys.back('start')
    await msg.answer(cnt, keyboard=key)


@bp.on.private_message(StateRule(AddOrder.AGENT))
async def verify_add_order(msg: Message):
    user = await msg.get_user()
    log.info(f'Called by {user.first_name} {user.last_name} ({user.id})')
    cnt = 'Подтвердите заказ:\n\n'
    m_payload = json.loads(msg.payload)
    agent_uid = m_payload.get('agent_uid', '')
    s_payload = msg.state_peer.payload
    async with DB_LOCK:
        agent = await Database.get_agent_by_uid(agent_uid)
    if not agent:
        log.warning(
            f'User #{msg.peer_id} selected unexistsing Agent UID: {agent_uid}'
        )
        await bp.state_dispenser.delete(msg.peer_id)
        await msg.answer(
            'ОШИБКА: Выбранный Агент не найден!',
            keyboard=keys.back('orders')
        )
        return
    await bp.state_dispenser.set(
        msg.peer_id, AddOrder.VERIFY, agent_uid=agent_uid,
        name=s_payload.get('name'), price=s_payload.get('price')
    )
    cnt += f'Цель: {s_payload.get("name")}\n'
    cnt += f'Цена: {s_payload.get("price")}\n'
    cnt += f'Агент: {agent.name}\n\n'
    cnt += 'Создать заказ?'
    await msg.answer(cnt, keyboard=keys.verify())


@bp.on.private_message(StateRule(AddOrder.VERIFY))
async def add_order_end(msg: Message):
    user = await msg.get_user()
    log.info(f'Called by {user.first_name} {user.last_name} ({user.id})')
    if msg.text in keys.YES_TEXTS:
        payload = msg.state_peer.payload
        async with DB_LOCK:
            order = await Database.add_order(
                payload.get('name'), payload.get('price'),
                payload.get('agent_uid')
            )
        cnt = f'Заказ #{order.uid} - создан!'
        key = keys.order_btn(order.uid)
    else:
        cnt = 'Операция отменена!'
        key = keys.back('orders')
    await bp.state_dispenser.delete(msg.peer_id)
    await msg.answer(cnt, keyboard=key)


@bp.on.private_message(FuncRule(lambda m: is_order_cmd(m.payload)))
async def order(msg: Message):
    user = await msg.get_user()
    log.info(f'Called by {user.first_name} {user.last_name} ({user.id})')
    order_uid = int(msg.get_payload_json().get('command', '#').split('#')[1])
    async with DB_LOCK:
        order = await Database.get_order_by_uid(order_uid)
    if order:
        await msg.answer(
            order.get_full_str(),
            keyboard=keys.order_ctrl(order.uid)
        )
    else:
        await msg.answer(
            f'ОШИБКА: Заказ #{order_uid} - не найден!',
            keyboard=keys.back('orders')
        )


@bp.on.private_message(payload={'command': 'orders_history'})
async def orders_history(msg: Message):
    user = await msg.get_user()
    log.info(f'Called by {user.first_name} {user.last_name} ({user.id})')
    async with DB_LOCK:
        orders = await Database.get_orders()
    if orders:
        ended_orders = filter(
            (lambda o: o.end_date),
            orders
        )
        cnt = 'Оплаченые заказы:\n\n'
        flag = False
        full_price = 0
        for ord in ended_orders:
            flag = True
            cnt += f'{ord.get_short_str()}\n'
            full_price += ord.price
        if not flag:
            cnt += 'Пока-что ни один заказ не был оплачен.'
        else:
            cnt += f'\nИтого: {full_price} руб.'
    else:
        cnt = 'Заказы отсутствуют.'
    await msg.answer(
        cnt, keyboard=keys.orders_history()
    )


@bp.on.private_message(FuncRule(lambda m: is_del_order_cmd(m.payload)))
async def del_order(msg: Message):
    user = await msg.get_user()
    log.info(f'Called by {user.first_name} {user.last_name} ({user.id})')
    order_uid = int(msg.get_payload_json().get('command', '#').split('#')[1])
    async with DB_LOCK:
        await Database.del_order(order_uid)
    await msg.answer(
        f'Заказ #{order_uid} - удалён!',
        keyboard=keys.back('orders')
    )


@bp.on.private_message(FuncRule(lambda m: is_end_order_cmd(m.payload)))
async def end_order(msg: Message):
    user = await msg.get_user()
    log.info(f'Called by {user.first_name} {user.last_name} ({user.id})')
    order_uid = int(msg.get_payload_json().get('command', '#').split('#')[1])
    async with DB_LOCK:
        await Database.end_order(order_uid)
    await msg.answer(
        f'Заказ #{order_uid} - оплачен!',
        keyboard=keys.back('orders')
    )


@bp.on.private_message(payload={'command': 'order_info'})
async def order_info_start(msg: Message):
    user = await msg.get_user()
    log.info(f'Called by {user.first_name} {user.last_name} ({user.id})')
    await bp.state_dispenser.set(msg.peer_id, OrderInfo.UID)
    await msg.answer(
        'Отправьте номер заказа',
        keyboard=keys.back('orders')
    )


@bp.on.private_message(StateRule(OrderInfo.UID))
async def end_order_info(msg: Message):
    user = await msg.get_user()
    log.info(f'Called by {user.first_name} {user.last_name} ({user.id})')
    if not msg.text.isdigit():
        await msg.answer(
            'ОШИБКА: Номер заказа должен быть цифровым!\nПовторите ввод.',
            keyboard=keys.back('orders')
        )
        return
    await bp.state_dispenser.delete(msg.peer_id)
    async with DB_LOCK:
        order = await Database.get_order_by_uid(int(msg.text))
    if order:
        await msg.answer(
            order.get_full_str(), keyboard=keys.back('orders')
        )
    else:
        await msg.answer(
            'ОШИБКА: Заказ не найден!',
            keyboard=keys.back('orders')
        )


@bp.on.message(payload={'command': 'orders_history_export'})
async def orders_history_export(msg: Message):
    user = await msg.get_user()
    log.info(f'Called by {user.first_name} {user.last_name} ({user.id})')
    wait_msg = await msg.answer('Файл готовится...')
    async with DB_LOCK:
        orders = await Database.get_orders()
    file = await make_export_file(orders)
    period = (
        orders[0].start_date,
        orders[::-1][0].start_date
    )
    file_name = f'Отчёт с {str(period[0])} по {str(period[1])}.xlsx'
    uploader = DocMessagesUploader(bp.api)
    doc = await uploader.upload(
        file_name, file, peer_id=msg.peer_id
    )
    await bp.api.messages.delete(
        wait_msg.message_id, peer_id=wait_msg.peer_id
    )
    await msg.answer(
        'Файл готов!', attachment=doc, keyboard=keys.start()
    )
