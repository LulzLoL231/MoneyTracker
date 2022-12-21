# -*- coding: utf-8 -*-
#
#  MoneyTracker cmds: Orders.
#  Created by LulzLoL231 at 12/11/22
#
from io import BytesIO
from dataclasses import dataclass

from openpyxl import Workbook
from openpyxl.styles import Font
from aiogram import types
from aiogram.dispatcher.storage import FSMContext
from aiogram.dispatcher.filters.state import State, StatesGroup

from config import check_admin
from runtimes import log, bot
from database.schemas import Order
from database.main import db as Database
from keyboards import Keyboards as keys


@dataclass
class OrderShortcut:
    command: str
    name: str
    price: str
    agent: str


class AddOrder(StatesGroup):
    NAME = State()
    PRICE = State()
    AGENT = State()
    VERIFY = State()


class OrderInfo(StatesGroup):
    UID = State()


class OrderPrice(StatesGroup):
    PRICE = State()


def is_digit(data: str) -> bool:
    '''Is digit in data?

    Args:
        data (str): Any string.

    Returns:
        bool: Boolean.
    '''
    try:
        int(data)
    except Exception:
        return False
    else:
        return True


async def make_export_file(orders: list[Order]) -> BytesIO:
    '''Makes XLSX file and return it.

    Args:
        orders (list[Order]): Array of Orders.

    Returns:
        BytesIO: XLSX file.
    '''
    period = (
        orders[0].start_date,
        orders[::-1][0].start_date
    )
    wb = Workbook()
    ws = wb.active
    ws.title = f'с {str(period[0])} по {str(period[1])}'
    ws['A1'] = '№ Заказа'
    ws['A1'].font = Font(bold=True)
    ws['B1'] = 'Цель'
    ws['B1'].font = Font(bold=True)
    ws['C1'] = 'Цена'
    ws['C1'].font = Font(bold=True)
    ws['D1'] = 'Агент'
    ws['D1'].font = Font(bold=True)
    ws['E1'] = 'Дата начала'
    ws['E1'].font = Font(bold=True)
    ws['F1'] = 'Дата оплаты'
    ws['F1'].font = Font(bold=True)
    rid = 2
    for order in orders:
        ws[f'A{rid}'] = order.uid
        ws[f'B{rid}'] = order.name
        if order.price:
            ws[f'C{rid}'] = order.price
        else:
            ws[f'C{rid}'] = 'Н/д'
        ws[f'D{rid}'] = order.agent.name
        ws[f'E{rid}'] = str(order.start_date)
        if order.end_date:
            ws[f'F{rid}'] = str(order.end_date)
        else:
            ws[f'F{rid}'] = 'Не оплачен'
        rid += 1
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@bot.message_handler(
    lambda m: m.text.lower().split(';')[0] == 'create_order',
    state='*'
)
@check_admin()
async def create_order_shortcut(msg: types.Message, state: FSMContext):
    log.info(f'Called by {msg.chat.mention} ({msg.chat.id})')
    if state:
        await state.finish()
    try:
        order = OrderShortcut(*msg.text.split(';'))
    except TypeError:
        log.warning(
            f'User #{msg.chat.id} use incorrect syntax for '
            'order creation shortcut!'
        )
        await msg.answer('<b>ОШИБКА:</b> Неверный синтаксис!')
        return
    agents = await Database.get_agents()
    agent_l = list(filter(
        lambda a: a.name.lower() == order.agent.lower(), agents
    ))
    if not agent_l:
        log.warning(
            f'User #{msg.chat.id} selected unexistsing '
            f'Agent name: {order.agent}'
        )
        await msg.answer('<b>ОШИБКА:</b> Выбранный Агент не найден!')
        return
    else:
        agent = agent_l[0]
    if not is_digit(order.price):
        log.warning(
            f'User #{msg.chat.id} use not digit in order price!'
        )
        await msg.answer('<b>ОШИБКА:</b> Цена заказа не число!')
        return
    await AddOrder.VERIFY.set()
    cnt = 'Подтвердите заказ:\n\n'
    cnt += f'Цель: {order.name}\n'
    cnt += f'Цена: {order.price}\n'
    cnt += f'Агент: {agent.name}\n\n'
    cnt += 'Создать заказ?'
    await msg.answer(cnt, reply_markup=keys.inline_verify_order())


@bot.callback_query_handler(lambda q: q.data == 'orders', state='*')
@check_admin()
async def query_orders(
    query: types.CallbackQuery, state: FSMContext, only_edit=False
):
    log.info(
        f'Called by {query.message.chat.mention} ({query.message.chat.id})'
    )
    if state:
        await state.finish()
    orders = await Database.get_orders()
    inprog_orders = list(filter(
        (lambda o: not o.end_date),
        orders
    ))
    if not inprog_orders:
        cnt = 'Все заказы выполнены. Добавьте новый!'
    elif not orders:
        cnt = 'Заказы отсутствуют. Добавьте новый!'
    else:
        cnt = 'Текущие заказы:\n\n'
        full_price = 0
        for ord in inprog_orders:
            cnt += f'{ord.get_short_str()}\n'
            if ord.price:
                full_price += ord.price
        cnt += f'\nИтого: {full_price} руб.'
    if not only_edit:
        await query.answer()
    await query.message.edit_text(cnt, reply_markup=keys.orders(
        inprog_orders, bool(orders)
    ))


@bot.callback_query_handler(
    lambda q: q.data == 'add_order', state='*'
)
@check_admin()
async def query_add_order(query: types.CallbackQuery, state: FSMContext):
    log.info(
        f'Called by {query.message.chat.mention} ({query.message.chat.id})'
    )
    await query.answer()
    await start_add_order(query.message, state, True)


@bot.message_handler(commands=['add_order'], state='*')
@check_admin()
async def start_add_order(msg: types.Message, state: FSMContext, query=False):
    log.info(f'Called by {msg.chat.mention} ({msg.chat.id})')
    cnt = '<b>Добавление заказа</b>\n\nВведите цель заказа.'
    if state:
        await state.finish()
    await AddOrder.NAME.set()
    if query:
        await msg.edit_text(cnt, reply_markup=keys.back('orders', 'Отмена'))
    else:
        await msg.answer(cnt, reply_markup=keys.back(text='Отмена'))


@bot.message_handler(content_types=types.ContentType.TEXT, state=AddOrder.NAME)
@check_admin()
async def add_order_name(msg: types.Message, state: FSMContext):
    log.info(f'Called by {msg.chat.mention} ({msg.chat.id})')
    await AddOrder.PRICE.set()
    await state.update_data(name=msg.text)
    await msg.answer(
        '<b>Добавление заказа</b>\n\nВведите цену заказа.',
        reply_markup=keys.no_price()
    )


@bot.message_handler(
    content_types=types.ContentType.TEXT, state=AddOrder.PRICE
)
@check_admin()
async def add_order_price(msg: types.Message, state: FSMContext):
    log.info(f'Called by {msg.chat.mention} ({msg.chat.id})')
    if msg.text != 'Нету цены' and not is_digit(msg.text):
        await msg.answer(
            'ОШИБКА: В качестве цены должно выступать цифровое значение!'
            '\nПовторите ввод.',
            reply_markup=keys.no_price()
        )
        return
    await AddOrder.AGENT.set()
    agents = await Database.get_agents()
    if agents:
        cnt = 'Выберите агента'
        key = keys.add_order_agents(agents)
        if msg.text == 'Нету цены':
            await state.update_data(price=msg.text)
        else:
            await state.update_data(price=int(msg.text))
    else:
        cnt = 'Нету доступных агентов. Сначала добавьте хотя-бы одного.'
        key = keys.back('agents', 'к Агентам')
        await state.finish()
    await msg.answer(cnt, reply_markup=key)


@bot.callback_query_handler(
    lambda q: q.data.startswith('agent_order'), state=AddOrder.AGENT
)
@check_admin()
async def verify_add_order(query: types.CallbackQuery, state: FSMContext):
    log.info(
        f'Called by {query.message.chat.mention} ({query.message.chat.id})'
    )
    cnt = '<b>Подтвердите заказ:</b>\n\n'
    agent_uid = int(query.data.split('#')[1])
    agent = await Database.get_agent_by_uid(agent_uid)
    await state.update_data(agent_uid=agent.uid)
    data = await state.get_data()
    if not agent:
        log.warning(
            f'User #{query.message.chat.id} selected unexistsing '
            f'Agent UID: {agent_uid}'
        )
        await state.finish()
        await query.answer()
        await query.message.edit_text(
            'ОШИБКА: Выбранный Агент не найден!',
            reply_markup=keys.back('orders')
        )
        return
    await AddOrder.VERIFY.set()
    cnt += f'Цель: {data.get("name")}\n'
    cnt += f'Цена: {data.get("price")}\n'
    cnt += f'Агент: {agent.name}\n\n'
    cnt += 'Создать заказ?'
    await query.answer()
    await query.message.edit_text(cnt, reply_markup=keys.inline_verify_order())


@bot.callback_query_handler(
    lambda q: q.data == 'verify_order', state=AddOrder.VERIFY
)
@check_admin()
async def add_order_end(query: types.CallbackQuery, state: FSMContext):
    log.info(
        f'Called by {query.message.chat.mention} ({query.message.chat.id})'
    )
    data = await state.get_data()
    if data.get('price') == 'Нету цены':
        price = None
    else:
        price = data.get('price')
    order = await Database.add_order(
        data.get('name'), data.get('agent_uid'), price
    )
    Database.get_order_by_uid.cache_clear()
    await state.finish()
    await query.answer()
    await query.message.edit_text(
        f'Заказ #{order.uid} - создан!',
        reply_markup=keys.order_btn(order.uid)
    )


@bot.callback_query_handler(lambda q: q.data.startswith('order'))
@check_admin()
async def order(query: types.CallbackQuery):
    log.info(
        f'Called by {query.message.chat.mention} ({query.message.chat.id})'
    )
    order_uid = int(query.data.split('#')[1])
    order = await Database.get_order_by_uid(order_uid)
    if order:
        await query.answer()
        await query.message.edit_text(
            order.get_full_str(), reply_markup=keys.order_ctrl(order)
        )
    else:
        await query.answer()
        await query.message.edit_text(
            f'<b>ОШИБКА:</b> Заказ #{order_uid} - не найден!',
            reply_markup=keys.back('orders')
        )


@bot.callback_query_handler(lambda q: q.data == 'ordres_history')
@check_admin()
async def ordres_history(query: types.CallbackQuery):
    log.info(
        f'Called by {query.message.chat.mention} ({query.message.chat.id})'
    )
    orders = await Database.get_orders()
    if orders:
        ended_orders = filter(
            (lambda o: o.end_date),
            orders
        )
        cnt = 'Оплаченые заказы:\n\n'
        flag = False
        full_price = 0
        for ord in ended_orders:
            flag = True
            cnt += f'{ord.get_short_str()}\n'
            if ord.price:
                full_price += ord.price
        if not flag:
            cnt += 'Пока-что ни один заказ не был оплачен.'
        else:
            cnt += f'\nИтого: {full_price} руб.'
    else:
        cnt = 'Заказы отсутствуют.'
    await query.answer()
    await query.message.edit_text(cnt, reply_markup=keys.order_history())


@bot.callback_query_handler(lambda q: q.data.startswith('del_order'))
@check_admin()
async def del_order(query: types.CallbackQuery):
    log.info(
        f'Called by {query.message.chat.mention} ({query.message.chat.id})'
    )
    order_uid = int(query.data.split('#')[1])
    await Database.del_order(order_uid)
    Database.get_order_by_uid.cache_clear()
    await query.answer(f'Заказ #{order_uid} - удалён!', True)
    await query_orders(query, bot.current_state(), True)


@bot.callback_query_handler(lambda q: q.data.startswith('end_order'))
@check_admin()
async def end_order(query: types.CallbackQuery):
    log.info(
        f'Called by {query.message.chat.mention} ({query.message.chat.id})'
    )
    order_uid = int(query.data.split('#')[1])
    await Database.end_order(order_uid)
    Database.get_order_by_uid.cache_clear()
    await query.answer(f'Заказ #{order_uid} - оплачен!', True)
    await query_orders(query, bot.current_state(), True)


@bot.callback_query_handler(lambda q: q.data.startswith('set_price_order'))
@check_admin()
async def start_set_price_order(query: types.CallbackQuery):
    log.info(
        f'Called by {query.message.chat.mention} ({query.message.chat.id})'
    )
    order_uid = int(query.data.split('#')[1])
    await OrderPrice.PRICE.set()
    state = bot.current_state()
    await state.update_data(order_uid=order_uid)
    cnt = f'<b>Установка цены для Заказа #{order_uid}</b>\n\n'
    cnt += 'Введите цену в рублях.'
    key = keys.back(f'order#{order_uid}', 'Отмена')
    await query.answer()
    await query.message.edit_text(cnt, reply_markup=key)


@bot.message_handler(content_types=types.ContentTypes.TEXT, state=OrderPrice.PRICE)
@check_admin()
async def end_set_price_order(msg: types.Message, state: FSMContext):
    log.info(f'Called by {msg.chat.mention} ({msg.chat.id})')
    data = await state.get_data()
    if not is_digit(msg.text):
        cnt = '<b>ОШИБКА:</b> Введено не число. Попробуйте ещё раз.'
        key = keys.back(f'order#{data.get("order_uid")}', 'Отмена')
        await msg.answer(cnt, reply_markup=key)
        return
    order = await Database.get_order_by_uid(data.get('order_uid'))
    if not order:
        await state.finish()
        cnt = f'<b>ОШИБКА:</b> Заказ {data.get("order_uid")} не найден!'
        key = keys.back('orders')
        await msg.answer(cnt, reply_markup=key)
        return
    if order.price:
        await state.finish()
        cnt = f'<b>ОШИБКА:</b> Цена для Заказа {data.get("order_uid")} уже установлена!'
        key = keys.back('orders')
        await msg.answer(cnt, reply_markup=key)
        return
    await state.finish()
    await Database.set_order_price(order.uid, int(msg.text))
    Database.get_order_by_uid.cache_clear()
    new_order = await Database.get_order_by_uid(data.get('order_uid'))
    if not new_order:
        await state.finish()
        cnt = f'<b>ОШИБКА:</b> Заказ {data.get("order_uid")} не найден! (ERR2)'
        key = keys.back('orders')
        await msg.answer(cnt, reply_markup=key)
        return
    cnt = f'<b>Цена установлена</b>\n\n{new_order.get_full_str()}'
    await msg.answer(cnt, reply_markup=keys.order_ctrl(new_order))


@bot.callback_query_handler(lambda q: q.data == 'info_order')
@check_admin()
async def order_info_start(query: types.CallbackQuery):
    log.info(
        f'Called by {query.message.chat.mention} ({query.message.chat.id})'
    )
    await OrderInfo.UID.set()
    await query.answer()
    await query.message.edit_text(
        '<b>Поиск заказа</b>\n\nОтправьте номер заказа.',
        reply_markup=keys.back(text='Отмена')
    )


@bot.message_handler(
    content_types=types.ContentType.TEXT, state=OrderInfo.UID
)
@check_admin()
async def end_order_info(msg: types.Message, state: FSMContext):
    log.info(f'Called by {msg.chat.mention} ({msg.chat.id})')
    if not msg.text.isdigit():
        await msg.answer(
            'ОШИБКА: Номер заказа должен быть цифровым!\nПовторите ввод.',
            reply_markup=keys.back(text='Отмена')
        )
        return
    await state.finish()
    order = await Database.get_order_by_uid(int(msg.text))
    if order:
        await msg.answer(
            order.get_full_str(), reply_markup=keys.order_ctrl(order)
        )
    else:
        await msg.answer(
            'ОШИБКА: Заказ не найден!',
            reply_markup=keys.back('orders')
        )


@bot.callback_query_handler(lambda q: q.data == 'export_order_history')
async def export_order_history(query: types.CallbackQuery):
    log.info(
        f'Called by {query.message.chat.mention} ({query.message.chat.id})'
    )
    await query.answer()
    await query.message.edit_text('<code>Файл готовится...</code>')
    await query.message.answer_chat_action(types.ChatActions.UPLOAD_DOCUMENT)
    orders = await Database.get_orders()
    file = await make_export_file(orders)
    period = (
        orders[0].start_date,
        orders[::-1][0].start_date
    )
    file_name = f'Отчёт с {str(period[0])} по {str(period[1])}.xlsx'
    await query.message.delete()
    await query.message.answer_document(
        types.InputFile(file, file_name),
        caption='Файл готов.'
    )
    await query_orders(query, bot.current_state(), True)
